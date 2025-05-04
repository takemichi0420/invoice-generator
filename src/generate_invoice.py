import argparse
import shutil
import pandas as pd
import datetime as dt
import glob
import openpyxl as op
import os
import unicodedata
from decimal import Decimal, ROUND_UP
from openpyxl.styles import PatternFill, Font


def detect_data_rows(ws, start_from=5):
    start_row = None
    end_row = 0
    for row in range(start_from, ws.max_row + 1):
        item = ws.cell(row=row, column=2).value
        if item is not None:
            if start_row is None:
                start_row = row
            end_row = row
    return start_row, end_row

def apply_styles_to_detail(ws, start_row, row_count):
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    font = Font(size=9)
    for idx in range(start_row, start_row + row_count):
        row_fill = gray_fill if (idx % 2 == 0) else white_fill
        for col in range(1, 6):
            cell = ws.cell(row=idx, column=col)
            cell.fill = row_fill
            cell.font = font
        ws.row_dimensions[idx].height = 13

def generate_invoice(client_name, delivery_date_cell, tax_rate=1.1):
    try:
 # パス設定
        input_dir = "/Users/matsushitatakemiti/Library/CloudStorage/OneDrive-個人用/02旅館その他 納品書・請求書"
        copied_input_path = "/Users/matsushitatakemiti/Library/CloudStorage/OneDrive-個人用/00請求書作成書類/000.xlsx"
        invoice_template_path = "/Users/matsushitatakemiti/Library/CloudStorage/OneDrive-個人用/00請求書作成書類/請求書.xlsx"
        output_dir = "/Users/matsushitatakemiti/Library/CloudStorage/OneDrive-個人用/00請求書作成書類/output"
        sales_log_path = "/Users/matsushitatakemiti/Library/CloudStorage/OneDrive-個人用/00請求書作成書類/売り上げ一覧（旅館その他）.xlsx"

        all_files = glob.glob(f"{input_dir}/*.xlsx")
        normalized_target = unicodedata.normalize('NFKC', client_name)
        matched_files = [
            f for f in all_files
            if normalized_target in unicodedata.normalize('NFKC', os.path.basename(f))
        ]

        if not matched_files:
            print(f"❗ ファイルが見つかりません: {client_name}")
            return

        for src_file in matched_files:
            print(f"✅ ファイル処理開始: {src_file}")

            shutil.copy(src_file, copied_input_path)
            shutil.copy(invoice_template_path, copied_input_path.replace("000", "001"))

            wb_input = op.load_workbook(copied_input_path, data_only=True)
            for sheet_name in ["原本", "請求書", "単価表"]:
                if sheet_name in wb_input.sheetnames: #"原本", "請求書", "単価表"が存在しなければスキップ
                    wb_input.remove(wb_input[sheet_name])
            wb_input.save(copied_input_path)
            wb_input.close()

            wb_input = op.load_workbook(copied_input_path, data_only=True)
            wb_invoice = op.load_workbook(copied_input_path.replace("000", "001"))
            ws_invoice_detail = wb_invoice["明細"]
            ws_invoice_cover = wb_invoice["請求書"]

            data_list = []
            for ws in wb_input.worksheets:
                delivery_date = ws[delivery_date_cell].value
                if not isinstance(delivery_date, (dt.date, dt.datetime)):
                    print(f"⚠️ シート「{ws.title}」のセル {delivery_date_cell} に有効な日付がありません: {delivery_date}")
                    continue
                delivery_date = delivery_date.date() if isinstance(delivery_date, dt.datetime) else delivery_date

                start_row, end_row = detect_data_rows(ws)
                if start_row is None:
                    continue

                for row_idx in range(start_row, end_row + 1):
                    item = ws.cell(row=row_idx, column=2).value
                    quantity = ws.cell(row=row_idx, column=3).value
                    unit_price = ws.cell(row=row_idx, column=4).value
                    amount = ws.cell(row=row_idx, column=5).value

                    if item:
                        data_list.append({
                            "納品日": delivery_date,
                            "品目": item,
                            "数量": quantity if item != "小計" else "",
                            "単価(¥)": unit_price or 0,
                            "金額(¥)": amount or 0
                        })

            if not data_list:
                print(f"❗ データが抽出できませんでした: {src_file}")
                continue

            df = pd.DataFrame(data_list)
            df = df.sort_values(by="納品日").reset_index(drop=True)

            # --- 型変換で数値列を明示的に揃える ---
            df["数量"] = pd.to_numeric(df["数量"], errors="coerce").fillna(0)
            df["単価(¥)"] = pd.to_numeric(df["単価(¥)"], errors="coerce").fillna(0)
            df["金額(¥)"] = pd.to_numeric(df["金額(¥)"], errors="coerce").fillna(0)

            subtotal = df["金額(¥)"].sum()
            df = pd.concat([
                df,
                pd.DataFrame([{
                    "納品日": None, "品目": "", "数量": "", "単価(¥)": "小計", "金額(¥)": subtotal
                }])
            ], ignore_index=True)

            for idx, row in df.iterrows():
                ws_invoice_detail.cell(row=2+idx, column=1, value=row["納品日"])
                ws_invoice_detail.cell(row=2+idx, column=2, value=row["品目"])
                ws_invoice_detail.cell(row=2+idx, column=3, value=row["数量"])
                ws_invoice_detail.cell(row=2+idx, column=4, value=row["単価(¥)"])
                ws_invoice_detail.cell(row=2+idx, column=5, value=row["金額(¥)"])

            apply_styles_to_detail(ws_invoice_detail, 2, len(df))

            for row in ws_invoice_detail.iter_rows(min_row=2, max_row=ws_invoice_detail.max_row):
                if isinstance(row[3].value, (int, float)):
                    row[3].number_format = '¥#,##0'
                if isinstance(row[4].value, (int, float)):
                    row[4].number_format = '¥#,##0'

            addressee = f"{client_name} 御中" if not client_name.endswith("御中") else client_name
            ws_invoice_detail.oddFooter.left.text = addressee
            ws_invoice_detail.oddFooter.right.text = "Page &P of &N"
            ws_invoice_cover["B3"] = addressee
            ws_invoice_cover["H21"] = subtotal

            today_str = dt.date.today().strftime("%Y_%m_%d")
            output_path = os.path.join(output_dir, f"{today_str}_{client_name}.xlsx")
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            wb_invoice.save(output_path)
            wb_invoice.close()
            wb_input.close()
            print(f"✅ 請求書作成完了: {output_path}")

            wb_sales = op.load_workbook(sales_log_path)
            ws_sales = wb_sales.active
            df_sales = pd.read_excel(sales_log_path, index_col=0)

            if client_name not in df_sales.index:
                print(f"⚠️ 売上一覧に取引先が見つかりません: {client_name}")
                wb_sales.close()
                continue

            supplier_row = df_sales.index.get_loc(client_name) + 2
            total_with_tax = (Decimal(str(subtotal)) * Decimal(str(tax_rate))).quantize(Decimal('1'), rounding=ROUND_UP)

            for col_idx in reversed(range(1, ws_sales.max_column + 1)):
                if ws_sales.cell(row=supplier_row, column=col_idx).value is not None:
                    ws_sales.cell(row=supplier_row, column=col_idx+1, value=float(total_with_tax))
                    break

            wb_sales.save(sales_log_path)
            wb_sales.close()
            print("✅ 売上一覧更新完了")

    except Exception as e:
        print(f"❌ エラー発生: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="請求書生成スクリプト")
    parser.add_argument("--client_name", required=True, help="取引先名（ファイル・請求書・売上一覧に共通）")
    parser.add_argument("--delivery_date_cell", required=True, help="納品日が入っているセル（例：E1）")
    parser.add_argument("--tax_rate", type=float, default=1.1, help="税率（例：1.1＝10%）")
    args = parser.parse_args()

    generate_invoice(
        client_name=args.client_name,
        delivery_date_cell=args.delivery_date_cell,
        tax_rate=args.tax_rate
    )
